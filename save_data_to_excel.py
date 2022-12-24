import string
import openpyxl
from typing import List
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from dataclasses import fields
from app_types import TGMessage
from database.controllers import message_controller

from config.db_config import EXCEL_TABLE_PATH, EXCEL_SHEET_NAME


class ExcelData:
    @staticmethod
    def __create_excel_file():
        """Creating file with data"""
        wb = openpyxl.Workbook()
        wb.save(EXCEL_TABLE_PATH)
        workbook = load_workbook(filename=EXCEL_TABLE_PATH)
        workbook.worksheets[0].title = EXCEL_SHEET_NAME
        wb.save(EXCEL_TABLE_PATH)
        return workbook

    def write_data_to_excel(self, data_filter: str):
        """Write data to excel"""
        workbook = self.__create_excel_file()
        src_data = message_controller.db_get_sorted_data(data_filter)
        if not src_data:
            return

        excel_column = 1
        excel_row = 1

        # write headers
        color_code = '9FC5E8'

        for field in fields(TGMessage):
            workbook[EXCEL_SHEET_NAME].cell(row=excel_row, column=excel_column).value = field.name
            workbook[EXCEL_SHEET_NAME].cell(row=excel_row, column=excel_column).fill = PatternFill(fgColor=color_code,
                                                                                                   fill_type='solid')
            excel_column += 1

        # increase column width
        for column in list(string.ascii_uppercase)[0:len(fields(TGMessage))]:
            workbook[EXCEL_SHEET_NAME].column_dimensions[column].width = 20

        # write values
        excel_row = 2
        for data in src_data:
            column_num = 0
            for field in fields(data):
                column_num += 1
                value = getattr(data, field.name)
                workbook[EXCEL_SHEET_NAME].cell(row=excel_row, column=column_num).value = value
                if column_num >= len(fields(data)):  # max users columns
                    break
            excel_row += 1

        workbook.save(EXCEL_TABLE_PATH)
        workbook.close()
        return True
